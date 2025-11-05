import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import re
import os

# --- Streamlit Page Setup ---
st.set_page_config(page_title="📦 SMW Box Contents Formatter", page_icon="📦", layout="wide")
st.title("📦 SMW Box Contents Formatter")
st.caption("Elegant Black Edition — Process, Pivot, and Format your Excel data instantly.")

# --- File Uploader ---
uploaded_file = st.file_uploader("📁 Select an Excel file", type=["xlsx", "xls"])

if uploaded_file:
    try:
        # Extract filename for output
        input_filename = uploaded_file.name
        base_name, ext = os.path.splitext(input_filename)
        output_filename = f"{base_name} formatted{ext}"

        # Read main sheet
        df = pd.read_excel(uploaded_file, header=10, engine="openpyxl")
        df.columns = df.columns.astype(str).str.strip()
    except Exception as e:
        st.error(f"❌ Error reading Excel file: {e}")
    else:
        required_columns = ["UPC", "Box X", "Sku Units"]
        missing_cols = [c for c in required_columns if c not in df.columns]

        if missing_cols:
            st.warning(f"⚠️ Missing columns: {', '.join(missing_cols)}")
        else:
            # --- Box Contents ---
            df_clean = df[required_columns].dropna(subset=["UPC", "Sku Units"])
            df_clean["UPC"] = (
                df_clean["UPC"]
                .astype(str)
                .str.replace(r"\.0$", "", regex=True)
                .str.replace("+", "", regex=False)
                .str.zfill(12)
            )
            df_clean["Sku Units"] = (
                pd.to_numeric(df_clean["Sku Units"], errors="coerce")
                .fillna(0)
                .astype(int)
            )
            df_clean.rename(columns={"Box X": "Box Number", "Sku Units": "Qty"}, inplace=True)

            # --- Pivot Table ---
            pivot_table = pd.pivot_table(
                df_clean,
                index="UPC",
                columns="Box Number",
                values="Qty",
                aggfunc="sum",
                fill_value=0,
            ).reset_index()
            pivot_table = pivot_table.replace(0, "")

            # --- Totals ---
            total_qty = df_clean["Qty"].sum()
            total_boxes = df_clean["Box Number"].nunique()

            # --- Extract Carton Weights ---
            carton_weights = []
            try:
                wb_input = load_workbook(uploaded_file, data_only=True)
                ws_page1 = wb_input["Page1_1"]
                for row in ws_page1.iter_rows(min_row=1, max_col=7):
                    cell = row[6]
                    if cell.font.bold and isinstance(cell.value, (int, float)):
                        carton_weights.append(cell.value)
                if carton_weights:
                    carton_weights = carton_weights[:-1]
            except Exception:
                carton_weights = []

            total_carton_weight = sum([w for w in carton_weights if isinstance(w, (int, float))])
            total_carton_weight_plus35 = total_carton_weight + 35

            # --- Extract Dimensions ---
            dimension_pattern = r"\b\d{1,3}\.\d{1,2}X\d{1,3}\.\d{1,2}X\d{1,3}\.\d{1,2}\b"
            dimension_data = []
            for _, row in df.iterrows():
                for col in df.columns:
                    val = str(row[col])
                    if re.match(dimension_pattern, val):
                        length, width, height = val.split("X")
                        dimension_data.append((float(length), float(width), float(height)))

            # --- Box Dimensions ---
            dim_df = pd.DataFrame()
            if dimension_data:
                dim_df = pd.DataFrame(dimension_data, columns=["Length", "Width", "Height"])
                dim_df.insert(0, "Box Number", range(1, len(dim_df) + 1))
                weights_column = carton_weights[: len(dim_df)] + [""] * max(0, len(dim_df) - len(carton_weights))
                dim_df.insert(1, "Carton Weight", weights_column)

            # --- Write to Excel ---
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_clean.to_excel(writer, sheet_name="Box Contents", index=False)
                pivot_table.to_excel(writer, sheet_name="Pivot Table", index=False)
                if not dim_df.empty:
                    dim_df.to_excel(writer, sheet_name="Box Dimensions", index=False)

            output.seek(0)
            wb = load_workbook(output)
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            header_font = Font(bold=True, size=14)
            align_center = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            def style_sheet(ws, keep_decimals=False, force_int_cols=[]):
                for row in ws.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = yellow_fill
                        cell.font = header_font
                        cell.alignment = align_center
                        cell.border = thin_border
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                    for col_idx, cell in enumerate(row, start=1):
                        cell.border = thin_border
                        cell.alignment = align_center
                        if isinstance(cell.value, (int, float)):
                            if col_idx in force_int_cols:
                                cell.number_format = "0"
                            elif keep_decimals:
                                cell.number_format = "0.00"
                            else:
                                cell.number_format = "0"
                for col_idx in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 18

            # --- Apply Formatting ---
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if sheet_name == "Box Dimensions":
                    style_sheet(ws, keep_decimals=True, force_int_cols=[1])
                elif sheet_name == "Pivot Table":
                    style_sheet(ws, keep_decimals=False)
                    # Reduce width to 1/4 excluding UPC
                    for col_idx in range(2, ws.max_column + 1):
                        current_width = ws.column_dimensions[get_column_letter(col_idx)].width
                        ws.column_dimensions[get_column_letter(col_idx)].width = current_width / 4
                    # Make UPC column slightly wider for clarity
                    ws.column_dimensions["A"].width = 25
                else:
                    style_sheet(ws, keep_decimals=False)

            # --- Totals for Box Contents ---
            ws_contents = wb["Box Contents"]
            total_row = ws_contents.max_row + 2
            ws_contents[f"A{total_row}"] = "Total Qty:"
            ws_contents[f"B{total_row}"] = total_qty
            ws_contents[f"A{total_row + 1}"] = "Total Boxes:"
            ws_contents[f"B{total_row + 1}"] = total_boxes
            for r in range(total_row, total_row + 2):
                for c in range(1, 3):
                    cell = ws_contents.cell(row=r, column=c)
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = align_center

            # --- Total Carton Weight ---
            if "Box Dimensions" in wb.sheetnames:
                ws_dim = wb["Box Dimensions"]
                carton_col = 2
                last_row = ws_dim.max_row
                total_weight = sum(
                    [
                        ws_dim.cell(row=r, column=carton_col).value
                        for r in range(2, last_row + 1)
                        if isinstance(ws_dim.cell(row=r, column=carton_col).value, (int, float))
                    ]
                )
                total_weight += 35
                total_row = last_row + 1
                ws_dim.cell(row=total_row, column=1, value="Total Carton Weight (+35):")
                ws_dim.cell(row=total_row, column=carton_col, value=total_weight)
                for col in [1, carton_col]:
                    cell = ws_dim.cell(row=total_row, column=col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                ws_dim.column_dimensions["A"].width = 30

            formatted_output = BytesIO()
            wb.save(formatted_output)
            formatted_output.seek(0)

            # --- Streamlit Download ---
            st.download_button(
                label=f"💾 Download {output_filename}",
                data=formatted_output,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # --- Streamlit Preview ---
            st.markdown("---")
            col1, col2, col3 = st.columns(3)
            col1.metric("📦 Total Qty", f"{total_qty}")
            col2.metric("🧱 Total Boxes", f"{total_boxes}")
            col3.metric("⚖️ Total Carton Weight + 35", f"{total_carton_weight_plus35}")

            st.subheader("✅ Box Contents")
            st.dataframe(df_clean)

            st.subheader("✅ Pivot Table")
            st.dataframe(pivot_table)

            if not dim_df.empty:
                st.subheader("✅ Box Dimensions")
                preview_dim = dim_df.copy()
                preview_dim.loc[len(preview_dim)] = [
                    "Total Carton Weight (+35)",
                    total_carton_weight_plus35,
                    "",
                    "",
                    "",
                ]
                st.dataframe(preview_dim)
